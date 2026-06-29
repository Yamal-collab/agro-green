from dotenv import load_dotenv
from pathlib import Path
load_dotenv(Path(__file__).parent / ".env")

import os, uuid, logging, bcrypt, jwt
from datetime import datetime, timezone, timedelta
from typing import Optional, List, Dict, Any
from fastapi import FastAPI, APIRouter, HTTPException, Depends, Request, Response
from fastapi.security import HTTPBearer, HTTPAuthorizationCredentials
from starlette.middleware.cors import CORSMiddleware
from motor.motor_asyncio import AsyncIOMotorClient
from pydantic import BaseModel, Field, EmailStr

client = AsyncIOMotorClient(os.environ["MONGO_URL"])
db = client[os.environ["DB_NAME"]]
JWT_SECRET = os.environ["JWT_SECRET"]
JWT_ALGO = "HS256"
TTL_MIN = 60 * 24

app = FastAPI(title="AgriBiz ERP")
api = APIRouter(prefix="/api")
bearer = HTTPBearer(auto_error=False)

def now_iso(): return datetime.now(timezone.utc).isoformat()
def new_id(): return str(uuid.uuid4())
def hp(p): return bcrypt.hashpw(p.encode(), bcrypt.gensalt()).decode()
def vp(p, h):
    try: return bcrypt.checkpw(p.encode(), h.encode())
    except: return False

def mk_token(uid, email):
    return jwt.encode({"sub": uid, "email": email, "exp": datetime.now(timezone.utc) + timedelta(minutes=TTL_MIN)}, JWT_SECRET, algorithm=JWT_ALGO)

async def get_user(request: Request, creds: Optional[HTTPAuthorizationCredentials] = Depends(bearer)):
    token = creds.credentials if creds else request.cookies.get("access_token")
    if not token: raise HTTPException(401, "Not authenticated")
    try: payload = jwt.decode(token, JWT_SECRET, algorithms=[JWT_ALGO])
    except: raise HTTPException(401, "Invalid token")
    u = await db.users.find_one({"id": payload["sub"]}, {"_id": 0, "password_hash": 0})
    if not u: raise HTTPException(401, "User not found")
    return u

async def next_seq(key: str) -> int:
    doc = await db.counters.find_one_and_update({"_id": key}, {"$inc": {"seq": 1}}, upsert=True, return_document=True)
    return doc["seq"]

async def gen_invoice(prefix: str) -> str:
    n = await next_seq(prefix)
    return f"{prefix}-{datetime.now().year}-{n:05d}"

async def fin_write(bu: int, type_: str, category: str, amount: float, date: str, ref_id: str = "", notes: str = "", source: str = ""):
    await db.finance_transactions.insert_one({
        "id": new_id(), "business_unit": bu, "type": type_, "category": category,
        "amount": round(amount, 2), "date": date, "source": source, "ref_id": ref_id,
        "notes": notes, "created_at": now_iso(),
    })

# ============ Models ============
class LoginIn(BaseModel):
    email: EmailStr; password: str

class CustomerIn(BaseModel):
    name: str; phone: str = ""; farm_name: str = ""; address: str = ""; gst: str = ""; business_units: List[int] = []

class SupplierIn(BaseModel):
    name: str; phone: str = ""; address: str = ""; gst: str = ""; business_unit: int = 1

class FeedItemIn(BaseModel):
    name: str; brand: str = ""; category: str = ""; unit: str = "kg"

class FeedPurchaseIn(BaseModel):
    supplier_id: str; feed_item_id: str; date: str; quantity: float; purchase_rate: float
    transport: float = 0.0; other: float = 0.0; payment_status: str = "pending"

class FeedSaleIn(BaseModel):
    customer_id: str; feed_item_id: str; date: str; quantity: float; unit_price: float
    transport: float = 0.0; discount: float = 0.0; payment_status: str = "pending"

class FeedTransferIn(BaseModel):
    feed_item_id: str; date: str; quantity: float; notes: str = ""

class EggPurchaseIn(BaseModel):
    supplier_id: str; date: str; quantity: int; rate: float; transport: float = 0.0
    incubation_start: str

class BatchUpdateIn(BaseModel):
    hatch_date: Optional[str] = None; hatched_chicks: Optional[int] = None
    dead_eggs: Optional[int] = None; status: Optional[str] = None; notes: Optional[str] = None

class BatchExpenseIn(BaseModel):
    batch_id: str; category: str; amount: float; date: str; notes: str = ""

class ChickSaleIn(BaseModel):
    batch_id: str; customer_id: str; date: str; quantity: int; unit_price: float
    transport: float = 0.0; discount: float = 0.0; payment_status: str = "pending"

class ChickTransferIn(BaseModel):
    batch_id: str; date: str; quantity: int; notes: str = ""

class FarmSaleIn(BaseModel):
    customer_id: str; date: str; quantity: int; unit_price: float
    transport: float = 0.0; discount: float = 0.0; payment_status: str = "pending"

class FarmExpenseIn(BaseModel):
    category: str; amount: float; date: str; notes: str = ""

class TankIn(BaseModel):
    name: str; capacity: float; current_liters: float = 0.0

class TankAddIn(BaseModel):
    tank_id: str; date: str; liters: float; source: str = ""; loading_charge: float = 0.0

class WaterSaleIn(BaseModel):
    customer_id: str; date: str; liters: float; rate: float; received: float = 0.0; notes: str = ""

class WaterExpenseIn(BaseModel):
    category: str; amount: float; date: str; notes: str = ""

class PaymentIn(BaseModel):
    customer_id: Optional[str] = None
    amount: float
    date: str
    method: str = "cash"
    notes: str = ""
    business_unit: int = 0  # 0 = cross-BU FIFO
    # Back-compat optional fields
    party_id: Optional[str] = None
    party_type: Optional[str] = "customer"

# ============ Auth ============
@api.post("/auth/login")
async def login(payload: LoginIn, response: Response):
    u = await db.users.find_one({"email": payload.email.lower()})
    if not u or not vp(payload.password, u["password_hash"]):
        raise HTTPException(401, "Invalid credentials")
    tok = mk_token(u["id"], u["email"])
    response.set_cookie("access_token", tok, httponly=True, secure=True, samesite="lax", max_age=TTL_MIN*60, path="/")
    return {"access_token": tok, "user": {"id": u["id"], "email": u["email"], "name": u["name"], "role": "admin"}}

@api.post("/auth/logout")
async def logout(response: Response, user=Depends(get_user)):
    response.delete_cookie("access_token", path="/")
    return {"ok": True}

@api.get("/auth/me")
async def me(user=Depends(get_user)): return user

# ============ Generic helpers ============
async def list_col(coll, q=None, limit=2000):
    return await coll.find(q or {}, {"_id": 0}).sort("created_at", -1).to_list(limit)

# ============ Customers (shared across BUs) ============
@api.get("/customers")
async def list_customers(user=Depends(get_user)): return await list_col(db.customers)

@api.post("/customers")
async def create_customer(p: CustomerIn, user=Depends(get_user)):
    d = {"id": new_id(), **p.model_dump(), "outstanding": 0.0, "status": "active", "created_at": now_iso()}
    await db.customers.insert_one(d); d.pop('_id', None); d.pop("_id", None); return d

# ============ Suppliers ============
@api.get("/suppliers")
async def list_suppliers(user=Depends(get_user)): return await list_col(db.suppliers)

@api.post("/suppliers")
async def create_supplier(p: SupplierIn, user=Depends(get_user)):
    d = {"id": new_id(), **p.model_dump(), "outstanding": 0.0, "created_at": now_iso()}
    await db.suppliers.insert_one(d); d.pop('_id', None); d.pop("_id", None); return d

# ============ BU1: Feed Items + Purchases + Sales + Transfer ============
@api.get("/feed/items")
async def feed_items(user=Depends(get_user)): return await list_col(db.feed_items)

@api.post("/feed/items")
async def feed_item_create(p: FeedItemIn, user=Depends(get_user)):
    d = {"id": new_id(), **p.model_dump(), "current_stock": 0.0, "weighted_avg_cost": 0.0, "created_at": now_iso()}
    await db.feed_items.insert_one(d); d.pop('_id', None); d.pop("_id", None); return d

@api.get("/feed/purchases")
async def feed_purchases(user=Depends(get_user)): return await list_col(db.feed_purchases)

@api.post("/feed/purchases")
async def feed_purchase(p: FeedPurchaseIn, user=Depends(get_user)):
    item = await db.feed_items.find_one({"id": p.feed_item_id})
    if not item: raise HTTPException(404, "Feed item not found")
    total = p.quantity * p.purchase_rate + p.transport + p.other
    eff_rate = total / p.quantity if p.quantity else 0
    cur_stock = item["current_stock"]; cur_avg = item["weighted_avg_cost"]
    new_stock = cur_stock + p.quantity
    new_avg = ((cur_stock * cur_avg) + (p.quantity * eff_rate)) / new_stock if new_stock else 0
    await db.feed_items.update_one({"id": p.feed_item_id}, {"$set": {"current_stock": new_stock, "weighted_avg_cost": round(new_avg, 4)}})
    d = {"id": new_id(), **p.model_dump(), "total_cost": round(total, 2), "created_at": now_iso()}
    await db.feed_purchases.insert_one(d); d.pop('_id', None)
    if p.payment_status != "paid":
        await db.suppliers.update_one({"id": p.supplier_id}, {"$inc": {"outstanding": total}})
    await fin_write(1, "expense", "Feed Purchase", total, p.date, d["id"], item["name"], "feed_purchase")
    return d

@api.get("/feed/sales")
async def feed_sales(user=Depends(get_user)): return await list_col(db.feed_sales)

@api.post("/feed/sales")
async def feed_sale(p: FeedSaleIn, user=Depends(get_user)):
    item = await db.feed_items.find_one({"id": p.feed_item_id})
    if not item or item["current_stock"] < p.quantity:
        raise HTTPException(400, "Insufficient stock")
    cust = await db.customers.find_one({"id": p.customer_id})
    if not cust: raise HTTPException(404, "Customer not found")
    total = p.quantity * p.unit_price + p.transport - p.discount
    cost = p.quantity * item["weighted_avg_cost"]
    inv = await gen_invoice("FE")
    amount_paid = round(total, 2) if p.payment_status == "paid" else 0.0
    balance_due = round(total - amount_paid, 2)
    d = {"id": new_id(), "invoice_no": inv, **p.model_dump(),
         "customer_name": cust["name"], "feed_name": item["name"],
         "total": round(total, 2), "amount_paid": amount_paid,
         "balance_due": balance_due, "business_unit": 1,
         "cost_basis": round(cost, 2), "created_at": now_iso()}
    await db.feed_sales.insert_one(d); d.pop('_id', None)
    await db.feed_items.update_one({"id": p.feed_item_id}, {"$inc": {"current_stock": -p.quantity}})
    if balance_due > 0:
        await db.customers.update_one({"id": p.customer_id}, {"$inc": {"outstanding": balance_due}})
    await fin_write(1, "income", "Feed Sale", total, p.date, d["id"], inv, "feed_sale")
    return d

@api.post("/feed/transfer")
async def feed_transfer(p: FeedTransferIn, user=Depends(get_user)):
    item = await db.feed_items.find_one({"id": p.feed_item_id})
    if not item or item["current_stock"] < p.quantity:
        raise HTTPException(400, "Insufficient stock")
    val = p.quantity * item["weighted_avg_cost"]
    await db.feed_items.update_one({"id": p.feed_item_id}, {"$inc": {"current_stock": -p.quantity}})
    d = {"id": new_id(), "type": "feed", "from_unit": 1, "to_unit": 3,
         "ref_item_id": p.feed_item_id, "item_name": item["name"],
         "quantity": p.quantity, "unit_cost": item["weighted_avg_cost"],
         "total_value": round(val, 2), "date": p.date, "notes": p.notes, "created_at": now_iso()}
    await db.internal_transfers.insert_one(d); d.pop('_id', None)
    await fin_write(1, "income", "Internal Transfer (Feed→Farm)", val, p.date, d["id"], "", "internal_transfer")
    await fin_write(3, "expense", "Feed Consumption", val, p.date, d["id"], item["name"], "internal_transfer")
    return d

# ============ BU2: Egg Purchases → Batches → Expenses → Chick Sales → Transfer ============
@api.get("/egg/purchases")
async def egg_purchases(user=Depends(get_user)): return await list_col(db.egg_purchases)

@api.post("/egg/purchases")
async def egg_purchase(p: EggPurchaseIn, user=Depends(get_user)):
    total = p.quantity * p.rate + p.transport
    batch_no = f"BATCH-{datetime.now().year}-{await next_seq('BATCH'):04d}"
    batch = {"id": new_id(), "batch_no": batch_no, "egg_qty": p.quantity,
             "incubation_start": p.incubation_start, "hatch_date": None,
             "hatched_chicks": 0, "dead_eggs": 0, "transferred": 0, "sold": 0,
             "status": "incubating", "notes": "", "created_at": now_iso()}
    await db.poultry_batches.insert_one(batch); batch.pop('_id', None)
    d = {"id": new_id(), "batch_id": batch["id"], **p.model_dump(),
         "total_cost": round(total, 2), "created_at": now_iso()}
    await db.egg_purchases.insert_one(d); d.pop('_id', None)
    if True:  # always log to supplier outstanding & finance
        await db.suppliers.update_one({"id": p.supplier_id}, {"$inc": {"outstanding": total}})
    await fin_write(2, "expense", "Egg Purchase", total, p.date, d["id"], batch_no, "egg_purchase")
    return {"purchase": d, "batch": batch}

@api.get("/hatchery/batches")
async def list_batches(user=Depends(get_user)): return await list_col(db.poultry_batches)

@api.patch("/hatchery/batches/{bid}")
async def update_batch(bid: str, p: BatchUpdateIn, user=Depends(get_user)):
    upd = {k: v for k, v in p.model_dump().items() if v is not None}
    if not upd: raise HTTPException(400, "No fields")
    await db.poultry_batches.update_one({"id": bid}, {"$set": upd})
    return await db.poultry_batches.find_one({"id": bid}, {"_id": 0})

@api.get("/hatchery/expenses")
async def batch_expenses(user=Depends(get_user)): return await list_col(db.batch_expenses)

@api.post("/hatchery/expenses")
async def add_batch_expense(p: BatchExpenseIn, user=Depends(get_user)):
    d = {"id": new_id(), **p.model_dump(), "created_at": now_iso()}
    await db.batch_expenses.insert_one(d); d.pop('_id', None)
    await fin_write(2, "expense", p.category, p.amount, p.date, d["id"], "", "batch_expense")
    return d

@api.get("/hatchery/sales")
async def chick_sales(user=Depends(get_user)): return await list_col(db.chick_sales)

@api.post("/hatchery/sales")
async def chick_sale(p: ChickSaleIn, user=Depends(get_user)):
    batch = await db.poultry_batches.find_one({"id": p.batch_id})
    cust = await db.customers.find_one({"id": p.customer_id})
    if not batch or not cust: raise HTTPException(404, "Batch/customer not found")
    avail = batch["hatched_chicks"] - batch.get("sold", 0) - batch.get("transferred", 0)
    if avail < p.quantity: raise HTTPException(400, f"Only {avail} chicks available")
    total = p.quantity * p.unit_price + p.transport - p.discount
    inv = await gen_invoice("CH")
    amount_paid = round(total, 2) if p.payment_status == "paid" else 0.0
    balance_due = round(total - amount_paid, 2)
    d = {"id": new_id(), "invoice_no": inv, **p.model_dump(),
         "customer_name": cust["name"], "batch_no": batch["batch_no"],
         "total": round(total, 2), "amount_paid": amount_paid,
         "balance_due": balance_due, "business_unit": 2, "created_at": now_iso()}
    await db.chick_sales.insert_one(d); d.pop('_id', None)
    await db.poultry_batches.update_one({"id": p.batch_id}, {"$inc": {"sold": p.quantity}})
    if balance_due > 0:
        await db.customers.update_one({"id": p.customer_id}, {"$inc": {"outstanding": balance_due}})
    await fin_write(2, "income", "Chick Sale", total, p.date, d["id"], inv, "chick_sale")
    return d

async def compute_batch_transfer_cost(batch_id: str) -> float:
    batch = await db.poultry_batches.find_one({"id": batch_id})
    egg = await db.egg_purchases.find_one({"batch_id": batch_id})
    exps = await db.batch_expenses.find({"batch_id": batch_id}).to_list(1000)
    inv = (egg["total_cost"] if egg else 0) + sum(e["amount"] for e in exps)
    n = max(batch.get("hatched_chicks", 0) or 1, 1)
    return round(inv / n, 4)

@api.post("/hatchery/transfer")
async def chick_transfer(p: ChickTransferIn, user=Depends(get_user)):
    batch = await db.poultry_batches.find_one({"id": p.batch_id})
    if not batch: raise HTTPException(404, "Batch not found")
    avail = batch["hatched_chicks"] - batch.get("sold", 0) - batch.get("transferred", 0)
    if avail < p.quantity: raise HTTPException(400, f"Only {avail} chicks available")
    unit_cost = await compute_batch_transfer_cost(p.batch_id)
    val = p.quantity * unit_cost
    await db.poultry_batches.update_one({"id": p.batch_id}, {"$inc": {"transferred": p.quantity}})
    d = {"id": new_id(), "type": "chicks", "from_unit": 2, "to_unit": 3,
         "ref_item_id": p.batch_id, "item_name": batch["batch_no"],
         "quantity": p.quantity, "unit_cost": unit_cost, "total_value": round(val, 2),
         "date": p.date, "notes": p.notes, "created_at": now_iso()}
    await db.internal_transfers.insert_one(d); d.pop('_id', None)
    await db.farm_stock.insert_one({"id": new_id(), "source_batch_id": p.batch_id,
         "batch_no": batch["batch_no"], "qty_received": p.quantity, "transfer_cost_per_bird": unit_cost,
         "current_count": p.quantity, "date": p.date, "created_at": now_iso()})
    await fin_write(2, "income", "Internal Transfer (Chicks→Farm)", val, p.date, d["id"], "", "internal_transfer")
    await fin_write(3, "expense", "Chick Purchase (internal)", val, p.date, d["id"], batch["batch_no"], "internal_transfer")
    return d

@api.get("/hatchery/batches/{bid}/pnl")
async def batch_pnl(bid: str, user=Depends(get_user)):
    batch = await db.poultry_batches.find_one({"id": bid}, {"_id": 0})
    if not batch: raise HTTPException(404, "Batch not found")
    egg = await db.egg_purchases.find_one({"batch_id": bid}, {"_id": 0})
    exps = await db.batch_expenses.find({"batch_id": bid}, {"_id": 0}).to_list(1000)
    sales = await db.chick_sales.find({"batch_id": bid}, {"_id": 0}).to_list(1000)
    transfers = await db.internal_transfers.find({"ref_item_id": bid, "type": "chicks"}, {"_id": 0}).to_list(1000)
    egg_cost = egg["total_cost"] if egg else 0
    exp_total = sum(e["amount"] for e in exps)
    investment = egg_cost + exp_total
    revenue = sum(s["total"] for s in sales)
    transfer_revenue = sum(t["total_value"] for t in transfers)
    return {"batch": batch, "egg_cost": egg_cost, "expenses_total": exp_total,
            "total_investment": investment, "sales_revenue": revenue,
            "transfer_revenue": transfer_revenue, "profit": round(revenue + transfer_revenue - investment, 2)}

# ============ BU3: Farm ============
@api.get("/farm/stock")
async def farm_stock(user=Depends(get_user)): return await list_col(db.farm_stock)

@api.get("/farm/sales")
async def farm_sales(user=Depends(get_user)): return await list_col(db.farm_sales)

@api.post("/farm/sales")
async def farm_sale(p: FarmSaleIn, user=Depends(get_user)):
    cust = await db.customers.find_one({"id": p.customer_id})
    if not cust: raise HTTPException(404, "Customer not found")
    total = p.quantity * p.unit_price + p.transport - p.discount
    inv = await gen_invoice("FA")
    amount_paid = round(total, 2) if p.payment_status == "paid" else 0.0
    balance_due = round(total - amount_paid, 2)
    d = {"id": new_id(), "invoice_no": inv, **p.model_dump(),
         "customer_name": cust["name"], "total": round(total, 2),
         "amount_paid": amount_paid, "balance_due": balance_due,
         "business_unit": 3, "created_at": now_iso()}
    await db.farm_sales.insert_one(d); d.pop('_id', None)
    # Decrement farm stock FIFO
    rem = p.quantity
    stocks = await db.farm_stock.find({"current_count": {"$gt": 0}}).sort("created_at", 1).to_list(1000)
    for s in stocks:
        if rem <= 0: break
        take = min(rem, s["current_count"])
        await db.farm_stock.update_one({"id": s["id"]}, {"$inc": {"current_count": -take}})
        rem -= take
    if balance_due > 0:
        await db.customers.update_one({"id": p.customer_id}, {"$inc": {"outstanding": balance_due}})
    await fin_write(3, "income", "Farm Sale", total, p.date, d["id"], inv, "farm_sale")
    return d

@api.get("/farm/expenses")
async def farm_expenses(user=Depends(get_user)): return await list_col(db.farm_expenses)

@api.post("/farm/expenses")
async def add_farm_expense(p: FarmExpenseIn, user=Depends(get_user)):
    d = {"id": new_id(), **p.model_dump(), "created_at": now_iso()}
    await db.farm_expenses.insert_one(d); d.pop('_id', None)
    await fin_write(3, "expense", p.category, p.amount, p.date, d["id"], "", "farm_expense")
    return d

# ============ BU4: Water ============
@api.get("/water/tanks")
async def tanks(user=Depends(get_user)): return await list_col(db.water_tanks)

@api.post("/water/tanks")
async def create_tank(p: TankIn, user=Depends(get_user)):
    d = {"id": new_id(), **p.model_dump(), "created_at": now_iso()}
    await db.water_tanks.insert_one(d); d.pop('_id', None); d.pop("_id", None); return d

@api.get("/water/tank-additions")
async def tank_additions(user=Depends(get_user)): return await list_col(db.water_tank_additions)

@api.post("/water/tank-additions")
async def add_to_tank(p: TankAddIn, user=Depends(get_user)):
    tank = await db.water_tanks.find_one({"id": p.tank_id})
    if not tank: raise HTTPException(404, "Tank not found")
    await db.water_tanks.update_one({"id": p.tank_id}, {"$inc": {"current_liters": p.liters}})
    d = {"id": new_id(), **p.model_dump(), "created_at": now_iso()}
    await db.water_tank_additions.insert_one(d); d.pop('_id', None)
    if p.loading_charge > 0:
        await fin_write(4, "expense", "Loading Charge", p.loading_charge, p.date, d["id"], p.source, "tank_addition")
    return d

@api.get("/water/sales")
async def water_sales(user=Depends(get_user)): return await list_col(db.water_sales)

@api.post("/water/sales")
async def water_sale(p: WaterSaleIn, user=Depends(get_user)):
    cust = await db.customers.find_one({"id": p.customer_id})
    if not cust: raise HTTPException(404, "Customer not found")
    total = p.liters * p.rate
    pending = max(0, total - p.received)
    pstatus = "paid" if pending == 0 else ("partial" if p.received > 0 else "pending")
    inv = await gen_invoice("WA")
    d = {"id": new_id(), **p.model_dump(),
         "invoice_no": inv,
         "customer_name": cust["name"], "total": round(total, 2),
         "pending": round(pending, 2),
         "amount_paid": round(p.received, 2),
         "balance_due": round(pending, 2),
         "payment_status": pstatus,
         "business_unit": 4,
         "created_at": now_iso()}
    await db.water_sales.insert_one(d); d.pop('_id', None)
    # Decrement first tank with stock
    tanks_ = await db.water_tanks.find({"current_liters": {"$gt": 0}}).sort("created_at", 1).to_list(50)
    rem = p.liters
    for t in tanks_:
        if rem <= 0: break
        take = min(rem, t["current_liters"])
        await db.water_tanks.update_one({"id": t["id"]}, {"$inc": {"current_liters": -take}})
        rem -= take
    if pending > 0:
        await db.customers.update_one({"id": p.customer_id}, {"$inc": {"outstanding": pending}})
    await fin_write(4, "income", "Water Sale", total, p.date, d["id"], cust["name"], "water_sale")
    return d

@api.get("/water/expenses")
async def water_expenses(user=Depends(get_user)): return await list_col(db.water_expenses)

@api.post("/water/expenses")
async def add_water_expense(p: WaterExpenseIn, user=Depends(get_user)):
    d = {"id": new_id(), **p.model_dump(), "created_at": now_iso()}
    await db.water_expenses.insert_one(d); d.pop('_id', None)
    await fin_write(4, "expense", p.category, p.amount, p.date, d["id"], "", "water_expense")
    return d

# ============ Payments (FIFO Customer Allocation) ============
SALE_COLLECTIONS = [
    ("feed_sales", 1, "FE"),
    ("chick_sales", 2, "CH"),
    ("farm_sales", 3, "FA"),
    ("water_sales", 4, "WA"),
]

def _derive_pstatus(amount_paid: float, total: float) -> str:
    if amount_paid <= 0: return "pending"
    if amount_paid + 0.0001 >= total: return "paid"
    return "partial"

async def _ensure_sale_payment_fields(coll_name: str, bu: int):
    """Backfill amount_paid / balance_due / business_unit / payment_status on legacy docs."""
    coll = db[coll_name]
    async for s in coll.find({"$or": [{"amount_paid": {"$exists": False}},
                                       {"balance_due": {"$exists": False}},
                                       {"business_unit": {"$exists": False}}]}):
        total = float(s.get("total", 0) or 0)
        if coll_name == "water_sales":
            paid = float(s.get("amount_paid", s.get("received", 0)) or 0)
            bal = float(s.get("balance_due", s.get("pending", max(0, total - paid))) or 0)
        else:
            pstatus = s.get("payment_status", "pending")
            paid = total if pstatus == "paid" else 0.0
            bal = round(total - paid, 2)
        upd = {"amount_paid": round(paid, 2), "balance_due": round(bal, 2),
               "business_unit": bu, "payment_status": _derive_pstatus(paid, total)}
        await coll.update_one({"id": s["id"]}, {"$set": upd})

async def _recompute_customer_outstanding(customer_id: str) -> float:
    total_bal = 0.0
    for coll_name, _bu, _pref in SALE_COLLECTIONS:
        rows = await db[coll_name].find(
            {"customer_id": customer_id, "balance_due": {"$gt": 0}},
            {"balance_due": 1, "_id": 0}
        ).to_list(5000)
        total_bal += sum(float(r.get("balance_due", 0) or 0) for r in rows)
    total_bal = round(total_bal, 2)
    await db.customers.update_one({"id": customer_id}, {"$set": {"outstanding": total_bal}})
    return total_bal

async def _customer_open_invoices(customer_id: str):
    """Return unpaid sales for a customer across all BUs, FIFO by created_at."""
    out = []
    for coll_name, bu, pref in SALE_COLLECTIONS:
        rows = await db[coll_name].find(
            {"customer_id": customer_id, "balance_due": {"$gt": 0}},
            {"_id": 0}
        ).to_list(5000)
        for r in rows:
            r["_collection"] = coll_name
            r["_bu"] = bu
            out.append(r)
    out.sort(key=lambda r: r.get("created_at", ""))
    return out

@api.get("/payments")
async def payments(user=Depends(get_user)): return await list_col(db.payments)

@api.post("/payments")
async def record_payment(p: PaymentIn, user=Depends(get_user)):
    # Resolve party
    party_type = (p.party_type or "customer").lower()
    if p.customer_id:
        party_type = "customer"
        party_id = p.customer_id
    elif p.party_id:
        party_id = p.party_id
    else:
        raise HTTPException(422, "customer_id is required")
    if p.amount is None or p.amount <= 0:
        raise HTTPException(422, "amount must be greater than zero")

    if party_type == "supplier":
        sup = await db.suppliers.find_one({"id": party_id})
        if not sup: raise HTTPException(404, "Supplier not found")
        d = {"id": new_id(), "customer_id": None, "party_id": party_id,
             "party_type": "supplier", "party_name": sup["name"],
             "amount": round(float(p.amount), 2), "date": p.date,
             "method": p.method, "notes": p.notes,
             "business_unit": p.business_unit or 1,
             "allocations": [], "created_at": now_iso()}
        await db.payments.insert_one(d); d.pop("_id", None)
        await db.suppliers.update_one({"id": party_id}, {"$inc": {"outstanding": -float(p.amount)}})
        await fin_write(d["business_unit"], "expense", "Supplier Payment (paid)", 0,
                        p.date, d["id"], sup["name"], "supplier_payment")
        return d

    # Customer payment — FIFO
    cust = await db.customers.find_one({"id": party_id})
    if not cust: raise HTTPException(404, "Customer not found")
    invoices = await _customer_open_invoices(party_id)
    remaining = round(float(p.amount), 2)
    allocations = []
    for inv in invoices:
        if remaining <= 0.0001: break
        bal = float(inv.get("balance_due", 0) or 0)
        if bal <= 0: continue
        applied = round(min(remaining, bal), 2)
        new_paid = round(float(inv.get("amount_paid", 0) or 0) + applied, 2)
        new_bal = round(bal - applied, 2)
        total = float(inv.get("total", 0) or 0)
        new_status = _derive_pstatus(new_paid, total)
        set_doc = {"amount_paid": new_paid, "balance_due": new_bal,
                   "payment_status": new_status}
        if inv["_collection"] == "water_sales":
            set_doc["pending"] = new_bal
        if inv["_collection"] == "water_sales":
            set_doc["received"] = new_paid
        await db[inv["_collection"]].update_one({"id": inv["id"]}, {"$set": set_doc})
        allocations.append({
            "sale_id": inv["id"], "sale_type": inv["_collection"],
            "business_unit": inv["_bu"], "invoice_no": inv.get("invoice_no", ""),
            "amount_applied": applied, "previous_balance": bal,
            "new_balance": new_bal, "new_status": new_status,
        })
        remaining = round(remaining - applied, 4)

    advance = round(remaining, 2) if remaining > 0.0001 else 0.0
    applied_total = round(float(p.amount) - advance, 2)

    # Persist payment record
    bu_for_fin = p.business_unit or (allocations[0]["business_unit"] if allocations else 1)
    d = {"id": new_id(), "customer_id": party_id, "party_id": party_id,
         "party_type": "customer", "party_name": cust["name"],
         "amount": round(float(p.amount), 2),
         "applied_amount": applied_total, "advance_amount": advance,
         "date": p.date, "method": p.method, "notes": p.notes,
         "business_unit": bu_for_fin, "allocations": allocations,
         "created_at": now_iso()}
    await db.payments.insert_one(d); d.pop("_id", None)

    # Recompute customer outstanding from authoritative source (sale balances)
    new_outstanding = await _recompute_customer_outstanding(party_id)
    d["customer_outstanding_after"] = new_outstanding

    # Write a finance income txn for the payment (cash received)
    await fin_write(bu_for_fin, "income", "Payment Received",
                    float(p.amount), p.date, d["id"], cust["name"], "payment")
    return d

# ============ Customer Details (invoices + payments) ============
@api.get("/customers/{cid}/details")
async def customer_details(cid: str, user=Depends(get_user)):
    cust = await db.customers.find_one({"id": cid}, {"_id": 0})
    if not cust: raise HTTPException(404, "Customer not found")
    invoices = []
    for coll_name, bu, pref in SALE_COLLECTIONS:
        rows = await db[coll_name].find({"customer_id": cid}, {"_id": 0}).sort("created_at", -1).to_list(5000)
        for r in rows:
            invoices.append({
                "id": r["id"],
                "invoice_no": r.get("invoice_no", ""),
                "date": r.get("date", ""),
                "business_unit": bu,
                "sale_type": coll_name,
                "total": float(r.get("total", 0) or 0),
                "amount_paid": float(r.get("amount_paid", 0) or 0),
                "balance_due": float(r.get("balance_due", 0) or 0),
                "payment_status": r.get("payment_status", "pending"),
                "created_at": r.get("created_at", ""),
            })
    invoices.sort(key=lambda x: x.get("created_at", ""), reverse=True)
    pays = await db.payments.find(
        {"$or": [{"customer_id": cid}, {"party_id": cid, "party_type": "customer"}]},
        {"_id": 0}
    ).sort("created_at", -1).to_list(2000)
    total_billed = sum(i["total"] for i in invoices)
    total_paid = sum(i["amount_paid"] for i in invoices)
    total_due = sum(i["balance_due"] for i in invoices)
    return {"customer": cust, "invoices": invoices, "payments": pays,
            "summary": {"total_billed": round(total_billed, 2),
                        "total_paid": round(total_paid, 2),
                        "total_due": round(total_due, 2)}}

# ============ Internal Transfers list ============
@api.get("/transfers")
async def transfers(user=Depends(get_user)): return await list_col(db.internal_transfers)

# ============ Finance / Reports ============
@api.get("/finance/transactions")
async def fin_tx(user=Depends(get_user), bu: Optional[int] = None, month: Optional[str] = None):
    q = {}
    if bu: q["business_unit"] = bu
    if month: q["date"] = {"$regex": f"^{month}"}
    return await list_col(db.finance_transactions, q, limit=5000)

@api.get("/finance/pnl")
async def pnl(user=Depends(get_user), bu: Optional[int] = None, month: Optional[str] = None):
    q = {}
    if bu: q["business_unit"] = bu
    if month: q["date"] = {"$regex": f"^{month}"}
    txs = await db.finance_transactions.find(q, {"_id": 0}).to_list(10000)
    income = sum(t["amount"] for t in txs if t["type"] == "income")
    expense = sum(t["amount"] for t in txs if t["type"] == "expense")
    by_cat: Dict[str, float] = {}
    for t in txs:
        if t["type"] == "expense":
            by_cat[t["category"]] = by_cat.get(t["category"], 0) + t["amount"]
    return {"income": round(income, 2), "expense": round(expense, 2),
            "profit": round(income - expense, 2), "expense_by_category": by_cat}

# ============ Reports (per-BU) ============
@api.get("/reports/outstanding")
async def reports_outstanding(user=Depends(get_user)):
    custs = await db.customers.find({"outstanding": {"$gt": 0}}, {"_id": 0}).to_list(5000)
    sups = await db.suppliers.find({"outstanding": {"$gt": 0}}, {"_id": 0}).to_list(5000)
    custs.sort(key=lambda c: c.get("outstanding", 0), reverse=True)
    sups.sort(key=lambda s: s.get("outstanding", 0), reverse=True)
    return {"customers": custs, "suppliers": sups}

@api.get("/reports/low-stock")
async def reports_low_stock(user=Depends(get_user)):
    items = await db.feed_items.find({"current_stock": {"$lt": 100}}, {"_id": 0}).to_list(2000)
    items.sort(key=lambda f: f.get("current_stock", 0))
    return items

# ============ Dashboard ============
@api.get("/dashboard/summary")
async def dashboard(user=Depends(get_user)):
    today = datetime.now(timezone.utc).strftime("%Y-%m-%d")
    month = datetime.now(timezone.utc).strftime("%Y-%m")
    txs = await db.finance_transactions.find({}, {"_id": 0}).to_list(20000)

    def agg(bu_filter=None):
        rev = sum(t["amount"] for t in txs if t["type"] == "income" and (bu_filter is None or t["business_unit"] == bu_filter) and t["date"].startswith(month))
        exp = sum(t["amount"] for t in txs if t["type"] == "expense" and (bu_filter is None or t["business_unit"] == bu_filter) and t["date"].startswith(month))
        return {"revenue": round(rev, 2), "expense": round(exp, 2), "profit": round(rev - exp, 2)}

    today_total = sum(t["amount"] for t in txs if t["type"] == "income" and t["date"].startswith(today))

    feed_items_data = await db.feed_items.find({}, {"_id": 0}).to_list(500)
    feed_stock_value = sum(f["current_stock"] * f["weighted_avg_cost"] for f in feed_items_data)
    low_stock = [f for f in feed_items_data if f["current_stock"] < 100]

    customers = await db.customers.find({}, {"_id": 0}).to_list(2000)
    outstanding = sum(c.get("outstanding", 0) for c in customers)
    top_customers = sorted(customers, key=lambda c: c.get("outstanding", 0), reverse=True)[:5]

    active_batches = await db.poultry_batches.count_documents({"status": {"$ne": "closed"}})
    farm_birds = sum(s["current_count"] for s in await db.farm_stock.find({}, {"_id": 0}).to_list(500))
    water_stock = sum(t["current_liters"] for t in await db.water_tanks.find({}, {"_id": 0}).to_list(50))

    # 6-month trend per BU
    trend = []
    for i in range(5, -1, -1):
        dt = datetime.now(timezone.utc).replace(day=1) - timedelta(days=i*30)
        mkey = dt.strftime("%Y-%m")
        row = {"month": mkey[5:]}
        for bu in [1, 2, 3, 4]:
            row[f"bu{bu}"] = round(sum(t["amount"] for t in txs if t["type"] == "income" and t["business_unit"] == bu and t["date"].startswith(mkey)), 2)
        trend.append(row)

    recent_tx = sorted(txs, key=lambda t: t["created_at"], reverse=True)[:10]

    return {
        "today_total_sales": round(today_total, 2),
        "combined": agg(),
        "outstanding": round(outstanding, 2),
        "bu1": {**agg(1), "feed_items": len(feed_items_data), "stock_value": round(feed_stock_value, 2), "low_stock_count": len(low_stock)},
        "bu2": {**agg(2), "active_batches": active_batches},
        "bu3": {**agg(3), "current_birds": farm_birds},
        "bu4": {**agg(4), "water_stock": round(water_stock, 2)},
        "top_customers": top_customers,
        "low_feed_stock": low_stock,
        "recent_transactions": recent_tx,
        "revenue_trend": trend,
    }

# ============ Invoice PDF / Print / Share ============
import io
from urllib.parse import quote
from fastapi.responses import StreamingResponse, HTMLResponse
from reportlab.lib.pagesizes import A5
from reportlab.lib import colors
from reportlab.lib.units import mm
from reportlab.pdfgen import canvas

BUSINESS_INFO = {
    "name": "AgriBiz ERP",
    "tagline": "Integrated Agri-Business Management",
    "address": "Poolapees, Malappuram, Kerala",
    "phone": "+91 98765 43210",
    "email": "info@agribiz.com",
    "gst": "33ABCDE1234F1Z5",
}

async def _load_sale(stype: str, sale_id: str):
    """Load sale, customer, and item-meta for given sale type."""
    stype = stype.lower()
    if stype == "feed":
        sale = await db.feed_sales.find_one({"id": sale_id}, {"_id": 0})
        if not sale: raise HTTPException(404, "Feed sale not found")
        cust = await db.customers.find_one({"id": sale["customer_id"]}, {"_id": 0}) or {}
        item = await db.feed_items.find_one({"id": sale["feed_item_id"]}, {"_id": 0}) or {}
        return {
            "type": "Feed Sale", "type_key": "feed", "sale": sale, "customer": cust,
            "item_name": sale.get("feed_name") or item.get("name") or "Feed",
            "unit": item.get("unit", "kg"),
        }
    if stype == "chick":
        sale = await db.chick_sales.find_one({"id": sale_id}, {"_id": 0})
        if not sale: raise HTTPException(404, "Chick sale not found")
        cust = await db.customers.find_one({"id": sale["customer_id"]}, {"_id": 0}) or {}
        return {
            "type": "Chick Sale", "type_key": "chick", "sale": sale, "customer": cust,
            "item_name": f"Chicks (Batch {sale.get('batch_no','')})",
            "unit": "nos",
        }
    if stype == "farm":
        sale = await db.farm_sales.find_one({"id": sale_id}, {"_id": 0})
        if not sale: raise HTTPException(404, "Farm sale not found")
        cust = await db.customers.find_one({"id": sale["customer_id"]}, {"_id": 0}) or {}
        return {
            "type": "Farm Sale", "type_key": "farm", "sale": sale, "customer": cust,
            "item_name": "Poultry Birds",
            "unit": "nos",
        }
    raise HTTPException(400, f"Invalid invoice type: {stype}")


def _build_invoice_pdf(ctx: dict) -> bytes:
    sale = ctx["sale"]; cust = ctx["customer"]
    quantity = float(sale.get("quantity", 0) or 0)
    unit_price = float(sale.get("unit_price", 0) or 0)
    transport = float(sale.get("transport", 0) or 0)
    discount = float(sale.get("discount", 0) or 0)
    subtotal = quantity * unit_price
    grand_total = float(sale.get("total", subtotal + transport - discount) or 0)

    buf = io.BytesIO()
    PAGE_W, PAGE_H = A5  # 148mm x 210mm
    c = canvas.Canvas(buf, pagesize=A5)
    M = 10 * mm

    # Header band
    c.setFillColor(colors.HexColor("#15803D"))
    c.rect(0, PAGE_H - 22 * mm, PAGE_W, 22 * mm, fill=1, stroke=0)
    c.setFillColor(colors.white)
    c.setFont("Helvetica-Bold", 18)
    c.drawString(M, PAGE_H - 11 * mm, BUSINESS_INFO["name"])
    c.setFont("Helvetica", 8)
    c.drawString(M, PAGE_H - 16 * mm, BUSINESS_INFO["tagline"])
    c.setFont("Helvetica-Bold", 11)
    c.drawRightString(PAGE_W - M, PAGE_H - 11 * mm, "TAX INVOICE")
    c.setFont("Helvetica", 7)
    c.drawRightString(PAGE_W - M, PAGE_H - 16 * mm, ctx["type"])

    y = PAGE_H - 28 * mm
    c.setFillColor(colors.black)

    # Business details + GST
    c.setFont("Helvetica", 7.5)
    c.drawString(M, y, BUSINESS_INFO["address"])
    y -= 3.5 * mm
    c.drawString(M, y, f"Phone: {BUSINESS_INFO['phone']}  |  Email: {BUSINESS_INFO['email']}")
    y -= 3.5 * mm
    c.setFont("Helvetica-Bold", 7.5)
    c.drawString(M, y, f"GSTIN: {BUSINESS_INFO['gst']}")
    y -= 5 * mm

    # Invoice meta box
    c.setStrokeColor(colors.HexColor("#E5E7EB"))
    c.setLineWidth(0.5)
    box_h = 16 * mm
    c.rect(M, y - box_h, PAGE_W - 2 * M, box_h, fill=0, stroke=1)
    c.setFont("Helvetica-Bold", 7)
    c.setFillColor(colors.HexColor("#6B7280"))
    c.drawString(M + 2 * mm, y - 4 * mm, "INVOICE NO.")
    c.drawString(M + 50 * mm, y - 4 * mm, "INVOICE DATE")
    c.drawString(M + 95 * mm, y - 4 * mm, "PAYMENT STATUS")
    c.setFillColor(colors.black)
    c.setFont("Helvetica-Bold", 10)
    c.drawString(M + 2 * mm, y - 9 * mm, str(sale.get("invoice_no", "—")))
    c.setFont("Helvetica", 9)
    c.drawString(M + 50 * mm, y - 9 * mm, str(sale.get("date", "—")))
    pstatus = str(sale.get("payment_status", "pending")).upper()
    pcolor = colors.HexColor("#15803D") if pstatus == "PAID" else (
        colors.HexColor("#C2410C") if pstatus == "PENDING" else colors.HexColor("#CA8A04"))
    c.setFillColor(pcolor)
    c.setFont("Helvetica-Bold", 9)
    c.drawString(M + 95 * mm, y - 9 * mm, pstatus)
    c.setFillColor(colors.black)
    y -= (box_h + 5 * mm)

    # Customer
    c.setFont("Helvetica-Bold", 8)
    c.setFillColor(colors.HexColor("#6B7280"))
    c.drawString(M, y, "BILL TO")
    y -= 4 * mm
    c.setFillColor(colors.black)
    c.setFont("Helvetica-Bold", 10)
    c.drawString(M, y, cust.get("name", "—"))
    y -= 4 * mm
    c.setFont("Helvetica", 8)
    if cust.get("farm_name"):
        c.drawString(M, y, cust["farm_name"]); y -= 3.5 * mm
    if cust.get("address"):
        c.drawString(M, y, cust["address"][:80]); y -= 3.5 * mm
    if cust.get("phone"):
        c.drawString(M, y, f"Phone: {cust['phone']}"); y -= 3.5 * mm
    if cust.get("gst"):
        c.drawString(M, y, f"GSTIN: {cust['gst']}"); y -= 3.5 * mm
    y -= 3 * mm

    # Items table header
    c.setFillColor(colors.HexColor("#F3F4F6"))
    c.rect(M, y - 6 * mm, PAGE_W - 2 * M, 6 * mm, fill=1, stroke=0)
    c.setFillColor(colors.HexColor("#374151"))
    c.setFont("Helvetica-Bold", 7.5)
    c.drawString(M + 2 * mm, y - 4 * mm, "DESCRIPTION")
    c.drawRightString(M + 78 * mm, y - 4 * mm, "QTY")
    c.drawRightString(M + 102 * mm, y - 4 * mm, "RATE")
    c.drawRightString(PAGE_W - M - 2 * mm, y - 4 * mm, "AMOUNT")
    y -= 6 * mm

    # Item row
    c.setFillColor(colors.black)
    c.setFont("Helvetica", 9)
    c.drawString(M + 2 * mm, y - 5 * mm, ctx["item_name"][:48])
    c.drawRightString(M + 78 * mm, y - 5 * mm, f"{quantity:g} {ctx['unit']}")
    c.drawRightString(M + 102 * mm, y - 5 * mm, f"Rs. {unit_price:,.2f}")
    c.drawRightString(PAGE_W - M - 2 * mm, y - 5 * mm, f"Rs. {subtotal:,.2f}")
    y -= 8 * mm

    # Totals breakdown
    c.setStrokeColor(colors.HexColor("#E5E7EB"))
    c.line(M, y, PAGE_W - M, y)
    y -= 4 * mm
    c.setFont("Helvetica", 8.5)
    def row(label, value, bold=False, color=None):
        nonlocal y
        c.setFont("Helvetica-Bold" if bold else "Helvetica", 9 if bold else 8.5)
        if color: c.setFillColor(color)
        c.drawRightString(M + 102 * mm, y, label)
        c.drawRightString(PAGE_W - M - 2 * mm, y, value)
        c.setFillColor(colors.black)
        y -= 4.5 * mm

    row("Subtotal", f"Rs. {subtotal:,.2f}")
    row("Transport", f"Rs. {transport:,.2f}")
    row("Discount", f"- Rs. {discount:,.2f}")
    extra_gap = 12 * mm   # change this value
    y -= extra_gap
      
    # Thin separator line above Grand Total
    c.setStrokeColor(colors.HexColor("#D1D5DB"))
    c.setLineWidth(0.8)
    c.line(M + 60 * mm, y + 6 * mm, PAGE_W - M, y + 6 * mm)

    # Black bold text
    c.setFillColor(colors.black)
    c.setFont("Helvetica-Bold", 11)

    text_y = y + 1.5 * mm

    c.drawRightString(M + 102 * mm, text_y, "GRAND TOTAL")
    c.drawRightString(
    PAGE_W - M - 2 * mm,
    text_y,
    f"Rs. {grand_total:,.2f}"
)

    c.setFillColor(colors.black)
    y -= 12 * mm
    
   
    # Thank-you + signature
    c.setFont("Helvetica-Oblique", 9)
    c.setFillColor(colors.HexColor("#15803D"))
    c.drawString(M, y, "Thank you for your business")
    c.setFillColor(colors.black)

    # Signature line bottom-right
    sig_y = 18 * mm
    c.setStrokeColor(colors.HexColor("#9CA3AF"))
    c.line(PAGE_W - M - 50 * mm, sig_y, PAGE_W - M, sig_y)
    c.setFont("Helvetica", 7.5)
    c.setFillColor(colors.HexColor("#6B7280"))
    c.drawRightString(PAGE_W - M, sig_y - 4 * mm, "Authorised Signatory")

    # Footer
    c.setFont("Helvetica", 6.5)
    c.drawString(M, 8 * mm, f"This is a computer-generated invoice  •  {BUSINESS_INFO['name']}")

    c.showPage()
    c.save()
    return buf.getvalue()


@api.get("/invoice/{stype}/{sale_id}/pdf")
async def invoice_pdf(stype: str, sale_id: str, user=Depends(get_user)):
    ctx = await _load_sale(stype, sale_id)
    pdf_bytes = _build_invoice_pdf(ctx)
    inv_no = ctx["sale"].get("invoice_no", sale_id)
    filename = f"Invoice-{inv_no}.pdf"
    return StreamingResponse(io.BytesIO(pdf_bytes), media_type="application/pdf",
        headers={"Content-Disposition": f'inline; filename="{filename}"'})


@api.get("/invoice/{stype}/{sale_id}/print", response_class=HTMLResponse)
async def invoice_print(stype: str, sale_id: str, request: Request, user=Depends(get_user)):
    ctx = await _load_sale(stype, sale_id)
    inv_no = ctx["sale"].get("invoice_no", sale_id)
    # PDF endpoint URL (relative path is fine inside an iframe)
    pdf_url = f"/api/invoice/{stype}/{sale_id}/pdf"
    html = f"""<!doctype html><html><head><meta charset="utf-8">
<title>Print Invoice {inv_no}</title>
<style>
  html,body{{margin:0;padding:0;height:100%;background:#f3f4f6;font-family:system-ui,sans-serif}}
  iframe{{width:100%;height:100vh;border:0;background:#fff}}
  .bar{{position:fixed;top:8px;right:8px;z-index:10}}
  .bar button{{background:#15803D;color:#fff;border:0;border-radius:6px;padding:8px 14px;font-weight:600;cursor:pointer}}
</style></head>
<body>
<div class="bar"><button onclick="doPrint()">Print Again</button></div>
<iframe id="pdf" src="{pdf_url}"></iframe>
<script>
  function doPrint() {{
    const f = document.getElementById('pdf');
    try {{ f.contentWindow.focus(); f.contentWindow.print(); }}
    catch(e) {{ window.print(); }}
  }}
  document.getElementById('pdf').addEventListener('load', function() {{
    setTimeout(doPrint, 600);
  }});
</script>
</body></html>"""
    return HTMLResponse(content=html)


@api.post("/invoice/{stype}/{sale_id}/share")
async def invoice_share(stype: str, sale_id: str, request: Request, user=Depends(get_user)):
    ctx = await _load_sale(stype, sale_id)
    sale = ctx["sale"]; cust = ctx["customer"]
    inv_no = sale.get("invoice_no", sale_id)
    total = float(sale.get("total", 0) or 0)
    base = str(request.base_url).rstrip("/")
    pdf_url = f"{base}/api/invoice/{stype}/{sale_id}/pdf"

    message = (
        f"Hello {cust.get('name','Customer')},\n\n"
        f"Please find your invoice from {BUSINESS_INFO['name']}.\n\n"
        f"Invoice No: {inv_no}\n"
        f"Date: {sale.get('date','')}\n"
        f"Amount: Rs. {total:,.2f}\n"
        f"Payment Status: {str(sale.get('payment_status','pending')).upper()}\n\n"
        f"Invoice PDF: {pdf_url}\n\n"
        f"Thank you for your business."
    )
    phone_raw = "".join(ch for ch in (cust.get("phone") or "") if ch.isdigit())
    if phone_raw and not phone_raw.startswith("91") and len(phone_raw) == 10:
        phone_raw = "91" + phone_raw
    wa_phone = phone_raw  # may be empty -> generic wa.me/?text=
    whatsapp_url = (f"https://wa.me/{wa_phone}?text={quote(message)}"
                    if wa_phone else f"https://wa.me/?text={quote(message)}")

    subject = f"Invoice {inv_no} from {BUSINESS_INFO['name']}"
    mailto_url = f"mailto:{quote(cust.get('email',''))}?subject={quote(subject)}&body={quote(message)}"

    return {
        "whatsapp_url": whatsapp_url,
        "mailto_url": mailto_url,
        "pdf_url": pdf_url,
        "invoice_no": inv_no,
    }


# ============ Customer Ledger / Statement ============
def _to_date(s: str) -> str:
    return (s or "")[:10]

def _ledger_filter(entry_date: str, dfrom: Optional[str], dto: Optional[str]) -> bool:
    if dfrom and entry_date < dfrom: return False
    if dto and entry_date > dto: return False
    return True

async def _customer_ledger(customer_id: str, dfrom: Optional[str] = None, dto: Optional[str] = None):
    cust = await db.customers.find_one({"id": customer_id}, {"_id": 0})
    if not cust: raise HTTPException(404, "Customer not found")

    invoices = []
    for coll_name, bu, _pref in SALE_COLLECTIONS:
        rows = await db[coll_name].find({"customer_id": customer_id}, {"_id": 0}).to_list(10000)
        for r in rows:
            invoices.append({
                "id": r["id"], "invoice_no": r.get("invoice_no", ""),
                "date": _to_date(r.get("date", "")),
                "business_unit": bu, "sale_type": coll_name,
                "total": float(r.get("total", 0) or 0),
                "amount_paid": float(r.get("amount_paid", 0) or 0),
                "balance_due": float(r.get("balance_due", 0) or 0),
                "payment_status": r.get("payment_status", "pending"),
                "created_at": r.get("created_at", ""),
            })
    payments = await db.payments.find(
        {"$or": [{"customer_id": customer_id}, {"party_id": customer_id, "party_type": "customer"}]},
        {"_id": 0}
    ).to_list(10000)
    for p in payments:
        p["date"] = _to_date(p.get("date", ""))

    # Opening balance = sum of debits/credits BEFORE dfrom
    opening = 0.0
    if dfrom:
        for inv in invoices:
            if inv["date"] < dfrom: opening += inv["total"]
        for p in payments:
            if p["date"] < dfrom: opening -= float(p.get("amount", 0) or 0)

    entries = []
    for inv in invoices:
        if _ledger_filter(inv["date"], dfrom, dto):
            entries.append({
                "date": inv["date"], "created_at": inv.get("created_at", ""),
                "kind": "sale",
                "description": f"Sale - {BU_LABEL_MAP.get(inv['business_unit'], '')} ({inv.get('invoice_no','')})",
                "debit": inv["total"], "credit": 0.0,
                "reference_id": inv["id"], "business_unit": inv["business_unit"],
                "invoice_no": inv.get("invoice_no", ""),
            })
    for p in payments:
        if _ledger_filter(p["date"], dfrom, dto):
            allocs = p.get("allocations", []) or []
            invs = ", ".join(a.get("invoice_no", "") for a in allocs if a.get("invoice_no"))
            desc = f"Payment via {p.get('method', 'cash')}"
            if invs: desc += f" → {invs}"
            entries.append({
                "date": p.get("date", ""), "created_at": p.get("created_at", ""),
                "kind": "payment",
                "description": desc,
                "debit": 0.0, "credit": float(p.get("amount", 0) or 0),
                "reference_id": p.get("id", ""),
                "business_unit": p.get("business_unit", 0),
                "method": p.get("method", "cash"),
                "notes": p.get("notes", ""),
                "allocations": allocs,
            })

    entries.sort(key=lambda e: (e.get("date", ""), e.get("created_at", "")))
    running = opening
    for e in entries:
        running = running + e["debit"] - e["credit"]
        e["running_balance"] = round(running, 2)

    total_debit = round(sum(e["debit"] for e in entries), 2)
    total_credit = round(sum(e["credit"] for e in entries), 2)
    closing = round(running, 2)
    total_billed = round(sum(i["total"] for i in invoices), 2)
    total_paid_all = round(sum(float(p.get("amount", 0) or 0) for p in payments), 2)

    last_purchase = max((i["date"] for i in invoices), default="")
    last_payment = max((p["date"] for p in payments), default="")

    return {
        "customer": cust,
        "from": dfrom, "to": dto,
        "opening_balance": round(opening, 2),
        "closing_balance": closing,
        "total_debit": total_debit,
        "total_credit": total_credit,
        "total_billed": total_billed,
        "total_paid": total_paid_all,
        "outstanding": float(cust.get("outstanding", 0) or 0),
        "last_purchase_date": last_purchase,
        "last_payment_date": last_payment,
        "entries": entries,
    }

BU_LABEL_MAP = {1: "Feed", 2: "Hatchery", 3: "Farm", 4: "Water"}

@api.get("/customers/{cid}/ledger")
async def customer_ledger(cid: str, dfrom: Optional[str] = None, dto: Optional[str] = None,
                          user=Depends(get_user)):
    return await _customer_ledger(cid, dfrom, dto)

def _build_statement_pdf(ctx: dict) -> bytes:
    from reportlab.lib.pagesizes import A4
    cust = ctx["customer"]
    buf = io.BytesIO()
    PAGE_W, PAGE_H = A4
    c = canvas.Canvas(buf, pagesize=A4)
    M = 15 * mm

    # Header band
    c.setFillColor(colors.HexColor("#15803D"))
    c.rect(0, PAGE_H - 22 * mm, PAGE_W, 22 * mm, fill=1, stroke=0)
    c.setFillColor(colors.white)
    c.setFont("Helvetica-Bold", 20)
    c.drawString(M, PAGE_H - 11 * mm, BUSINESS_INFO["name"])
    c.setFont("Helvetica", 9)
    c.drawString(M, PAGE_H - 17 * mm, BUSINESS_INFO["tagline"])
    c.setFont("Helvetica-Bold", 13)
    c.drawRightString(PAGE_W - M, PAGE_H - 11 * mm, "ACCOUNT STATEMENT")
    c.setFont("Helvetica", 8)
    period_line = ""
    if ctx.get("from") or ctx.get("to"):
        period_line = f"Period: {ctx.get('from') or 'Start'} → {ctx.get('to') or 'Today'}"
    else:
        period_line = "Period: All transactions"
    c.drawRightString(PAGE_W - M, PAGE_H - 17 * mm, period_line)

    y = PAGE_H - 28 * mm
    c.setFillColor(colors.black)
    c.setFont("Helvetica", 8)
    c.drawString(M, y, BUSINESS_INFO["address"])
    c.drawString(M, y - 4 * mm, f"Phone: {BUSINESS_INFO['phone']}  |  Email: {BUSINESS_INFO['email']}")
    c.setFont("Helvetica-Bold", 8)
    c.drawString(M, y - 8 * mm, f"GSTIN: {BUSINESS_INFO['gst']}")

    # Customer box
    y -= 16 * mm
    c.setStrokeColor(colors.HexColor("#E5E7EB"))
    c.rect(M, y - 22 * mm, PAGE_W - 2 * M, 22 * mm, fill=0, stroke=1)
    c.setFont("Helvetica-Bold", 8)
    c.setFillColor(colors.HexColor("#6B7280"))
    c.drawString(M + 3 * mm, y - 4 * mm, "BILL TO")
    c.setFillColor(colors.black)
    c.setFont("Helvetica-Bold", 12)
    c.drawString(M + 3 * mm, y - 9 * mm, cust.get("name", "—"))
    c.setFont("Helvetica", 9)
    if cust.get("farm_name") or cust.get("business_name"):
        c.drawString(M + 3 * mm, y - 13 * mm, cust.get("farm_name") or cust.get("business_name") or "")
    if cust.get("address"):
        c.drawString(M + 3 * mm, y - 17 * mm, (cust.get("address") or "")[:80])
    contact = []
    if cust.get("phone"): contact.append(f"Phone: {cust['phone']}")
    if cust.get("gst"): contact.append(f"GSTIN: {cust['gst']}")
    if contact:
        c.drawString(M + 3 * mm, y - 21 * mm, " · ".join(contact))

    y -= 28 * mm

    # Summary row
    box_w = (PAGE_W - 2 * M - 3 * 3 * mm) / 4
    summaries = [
        ("Opening Balance", f"Rs. {ctx['opening_balance']:,.2f}", colors.HexColor("#374151")),
        ("Total Purchases", f"Rs. {ctx['total_debit']:,.2f}", colors.HexColor("#C2410C")),
        ("Total Payments", f"Rs. {ctx['total_credit']:,.2f}", colors.HexColor("#15803D")),
        ("Closing Balance", f"Rs. {ctx['closing_balance']:,.2f}", colors.HexColor("#0F172A")),
    ]
    x = M
    for label, val, col in summaries:
        c.setStrokeColor(colors.HexColor("#E5E7EB"))
        c.rect(x, y - 14 * mm, box_w, 14 * mm, fill=0, stroke=1)
        c.setFont("Helvetica", 7)
        c.setFillColor(colors.HexColor("#6B7280"))
        c.drawString(x + 2 * mm, y - 4 * mm, label.upper())
        c.setFont("Helvetica-Bold", 11)
        c.setFillColor(col)
        c.drawString(x + 2 * mm, y - 11 * mm, val)
        x += box_w + 3 * mm
    c.setFillColor(colors.black)
    y -= 20 * mm

    # Ledger table
    c.setFont("Helvetica-Bold", 10)
    c.drawString(M, y, "LEDGER")
    y -= 5 * mm
    headers = ["Date", "Description", "Debit", "Credit", "Balance"]
    cols_x = [M, M + 22 * mm, M + 115 * mm, M + 140 * mm, M + 165 * mm]
    cols_w = [22 * mm, 93 * mm, 25 * mm, 25 * mm, PAGE_W - M - (M + 165 * mm)]

    def draw_header():
        c.setFillColor(colors.HexColor("#F3F4F6"))
        c.rect(M, y - 5 * mm, PAGE_W - 2 * M, 5 * mm, fill=1, stroke=0)
        c.setFillColor(colors.HexColor("#374151"))
        c.setFont("Helvetica-Bold", 7.5)
        c.drawString(cols_x[0] + 1 * mm, y - 3.5 * mm, headers[0])
        c.drawString(cols_x[1] + 1 * mm, y - 3.5 * mm, headers[1])
        c.drawRightString(cols_x[2] + cols_w[2] - 1 * mm, y - 3.5 * mm, headers[2])
        c.drawRightString(cols_x[3] + cols_w[3] - 1 * mm, y - 3.5 * mm, headers[3])
        c.drawRightString(PAGE_W - M - 1 * mm, y - 3.5 * mm, headers[4])
        c.setFillColor(colors.black)

    draw_header()
    y -= 6 * mm

    # Opening balance row
    c.setFont("Helvetica-Oblique", 8)
    c.drawString(cols_x[0] + 1 * mm, y, ctx.get("from") or "—")
    c.drawString(cols_x[1] + 1 * mm, y, "Opening Balance")
    c.drawRightString(PAGE_W - M - 1 * mm, y, f"Rs. {ctx['opening_balance']:,.2f}")
    y -= 5 * mm

    c.setFont("Helvetica", 8)
    for e in ctx["entries"]:
        if y < 25 * mm:
            c.showPage()
            y = PAGE_H - M
            draw_header()
            y -= 6 * mm
        c.drawString(cols_x[0] + 1 * mm, y, e["date"] or "—")
        desc = e["description"][:60]
        c.drawString(cols_x[1] + 1 * mm, y, desc)
        if e["debit"] > 0:
            c.setFillColor(colors.HexColor("#C2410C"))
            c.drawRightString(cols_x[2] + cols_w[2] - 1 * mm, y, f"{e['debit']:,.2f}")
            c.setFillColor(colors.black)
        if e["credit"] > 0:
            c.setFillColor(colors.HexColor("#15803D"))
            c.drawRightString(cols_x[3] + cols_w[3] - 1 * mm, y, f"{e['credit']:,.2f}")
            c.setFillColor(colors.black)
        c.drawRightString(PAGE_W - M - 1 * mm, y, f"{e['running_balance']:,.2f}")
        y -= 4.5 * mm

    # Closing
    if y < 30 * mm:
        c.showPage(); y = PAGE_H - M
    y -= 4 * mm
    c.setStrokeColor(colors.HexColor("#E5E7EB"))
    c.line(M, y, PAGE_W - M, y)
    y -= 5 * mm
    c.setFont("Helvetica-Bold", 9)
    c.drawString(cols_x[1] + 1 * mm, y, "CLOSING BALANCE")
    c.setFillColor(colors.HexColor("#15803D") if ctx['closing_balance'] <= 0 else colors.HexColor("#C2410C"))
    c.drawRightString(PAGE_W - M - 1 * mm, y, f"Rs. {ctx['closing_balance']:,.2f}")
    c.setFillColor(colors.black)

    # Footer
    y_footer = 15 * mm
    c.setFont("Helvetica-Oblique", 8)
    c.setFillColor(colors.HexColor("#15803D"))
    c.drawString(M, y_footer + 5 * mm, "Thank you for your business")
    c.setFillColor(colors.HexColor("#6B7280"))
    c.setFont("Helvetica", 6.5)
    c.drawString(M, y_footer, f"This is a computer-generated statement  •  {BUSINESS_INFO['name']}")
    c.setStrokeColor(colors.HexColor("#9CA3AF"))
    c.line(PAGE_W - M - 50 * mm, y_footer + 5 * mm, PAGE_W - M, y_footer + 5 * mm)
    c.setFont("Helvetica", 7)
    c.setFillColor(colors.HexColor("#6B7280"))
    c.drawRightString(PAGE_W - M, y_footer + 1 * mm, "Authorised Signatory")

    c.showPage()
    c.save()
    return buf.getvalue()


@api.get("/customers/{cid}/statement/pdf")
async def customer_statement_pdf(cid: str, dfrom: Optional[str] = None, dto: Optional[str] = None,
                                 user=Depends(get_user)):
    ctx = await _customer_ledger(cid, dfrom, dto)
    pdf_bytes = _build_statement_pdf(ctx)
    name = (ctx["customer"].get("name") or "customer").replace(" ", "_")
    fname = f"Statement-{name}.pdf"
    return StreamingResponse(io.BytesIO(pdf_bytes), media_type="application/pdf",
        headers={"Content-Disposition": f'inline; filename="{fname}"'})


@api.get("/customers/{cid}/statement/print", response_class=HTMLResponse)
async def customer_statement_print(cid: str, request: Request,
                                   dfrom: Optional[str] = None, dto: Optional[str] = None,
                                   user=Depends(get_user)):
    qs = ""
    if dfrom or dto:
        parts = []
        if dfrom: parts.append(f"dfrom={dfrom}")
        if dto: parts.append(f"dto={dto}")
        qs = "?" + "&".join(parts)
    pdf_url = f"/api/customers/{cid}/statement/pdf{qs}"
    html = f"""<!doctype html><html><head><meta charset="utf-8"><title>Print Statement</title>
<style>html,body{{margin:0;padding:0;height:100%;background:#f3f4f6;font-family:system-ui}}
iframe{{width:100%;height:100vh;border:0;background:#fff}}
.bar{{position:fixed;top:8px;right:8px;z-index:10}}
.bar button{{background:#15803D;color:#fff;border:0;border-radius:6px;padding:8px 14px;font-weight:600;cursor:pointer}}
</style></head><body>
<div class="bar"><button onclick="doPrint()">Print Again</button></div>
<iframe id="pdf" src="{pdf_url}"></iframe>
<script>
function doPrint(){{const f=document.getElementById('pdf');try{{f.contentWindow.focus();f.contentWindow.print();}}catch(e){{window.print();}}}}
document.getElementById('pdf').addEventListener('load',function(){{setTimeout(doPrint,600);}});
</script></body></html>"""
    return HTMLResponse(content=html)


@api.post("/customers/{cid}/statement/share")
async def customer_statement_share(cid: str, request: Request,
                                   dfrom: Optional[str] = None, dto: Optional[str] = None,
                                   user=Depends(get_user)):
    ctx = await _customer_ledger(cid, dfrom, dto)
    cust = ctx["customer"]
    base = str(request.base_url).rstrip("/")
    qs = ""
    if dfrom or dto:
        parts = []
        if dfrom: parts.append(f"dfrom={dfrom}")
        if dto: parts.append(f"dto={dto}")
        qs = "?" + "&".join(parts)
    pdf_url = f"{base}/api/customers/{cid}/statement/pdf{qs}"
    msg = (
        f"Hello {cust.get('name','Customer')},\n\n"
        f"Please find your account statement from {BUSINESS_INFO['name']}.\n\n"
        f"Period: {ctx.get('from') or 'Start'} → {ctx.get('to') or 'Today'}\n"
        f"Total Purchases: Rs. {ctx['total_debit']:,.2f}\n"
        f"Total Payments: Rs. {ctx['total_credit']:,.2f}\n"
        f"Outstanding Balance: Rs. {ctx['closing_balance']:,.2f}\n\n"
        f"Statement PDF: {pdf_url}\n\nThank you for your business."
    )
    phone_raw = "".join(ch for ch in (cust.get("phone") or "") if ch.isdigit())
    if phone_raw and not phone_raw.startswith("91") and len(phone_raw) == 10:
        phone_raw = "91" + phone_raw
    whatsapp_url = (f"https://wa.me/{phone_raw}?text={quote(msg)}"
                    if phone_raw else f"https://wa.me/?text={quote(msg)}")
    subject = f"Account Statement — {cust.get('name','Customer')}"
    mailto_url = f"mailto:{quote(cust.get('email',''))}?subject={quote(subject)}&body={quote(msg)}"
    return {"whatsapp_url": whatsapp_url, "mailto_url": mailto_url, "pdf_url": pdf_url}


# ============ Reports ============
def _within(s: dict, dfrom: Optional[str], dto: Optional[str], key: str = "date") -> bool:
    d = (s.get(key) or "")[:10]
    if dfrom and d < dfrom: return False
    if dto and d > dto: return False
    return True

@api.get("/reports/sales")
async def report_sales(dfrom: Optional[str] = None, dto: Optional[str] = None,
                       customer_id: Optional[str] = None, business_unit: Optional[int] = None,
                       user=Depends(get_user)):
    rows = []
    by_bu = {1: 0, 2: 0, 3: 0, 4: 0}
    by_status = {"paid": 0, "partial": 0, "pending": 0}
    for coll_name, bu, _pref in SALE_COLLECTIONS:
        if business_unit and business_unit != bu: continue
        q = {}
        if customer_id: q["customer_id"] = customer_id
        sales = await db[coll_name].find(q, {"_id": 0}).to_list(20000)
        for s in sales:
            if not _within(s, dfrom, dto): continue
            row = {
                "id": s["id"],
                "invoice_no": s.get("invoice_no", ""),
                "date": _to_date(s.get("date", "")),
                "customer_id": s.get("customer_id", ""),
                "customer_name": s.get("customer_name", ""),
                "business_unit": bu, "sale_type": coll_name,
                "total": float(s.get("total", 0) or 0),
                "amount_paid": float(s.get("amount_paid", 0) or 0),
                "balance_due": float(s.get("balance_due", 0) or 0),
                "payment_status": s.get("payment_status", "pending"),
            }
            rows.append(row)
            by_bu[bu] += row["total"]
            ps = row["payment_status"]
            if ps in by_status: by_status[ps] += row["total"]
    rows.sort(key=lambda r: r["date"], reverse=True)
    total = sum(r["total"] for r in rows)
    paid = sum(r["amount_paid"] for r in rows)
    due = sum(r["balance_due"] for r in rows)
    return {"rows": rows, "summary": {"count": len(rows),
            "total": round(total, 2), "paid": round(paid, 2), "due": round(due, 2),
            "by_business_unit": {k: round(v, 2) for k, v in by_bu.items()},
            "by_status": {k: round(v, 2) for k, v in by_status.items()}}}


@api.get("/reports/purchases")
async def report_purchases(dfrom: Optional[str] = None, dto: Optional[str] = None,
                           supplier_id: Optional[str] = None, business_unit: Optional[int] = None,
                           user=Depends(get_user)):
    rows = []
    pcollections = [
        ("feed_purchases", 1, "feed_item_id"),
        ("egg_purchases", 2, None),
    ]
    for coll_name, bu, _ in pcollections:
        if business_unit and business_unit != bu: continue
        q = {}
        if supplier_id: q["supplier_id"] = supplier_id
        purs = await db[coll_name].find(q, {"_id": 0}).to_list(20000)
        for p in purs:
            if not _within(p, dfrom, dto): continue
            total = float(p.get("total", 0) or 0)
            if not total:
                total = float(p.get("quantity", 0) or 0) * float(p.get("purchase_rate", p.get("rate", 0)) or 0) + float(p.get("transport", 0) or 0)
            sup = await db.suppliers.find_one({"id": p.get("supplier_id")}, {"name": 1, "_id": 0}) or {}
            rows.append({
                "id": p["id"], "date": _to_date(p.get("date", "")),
                "business_unit": bu, "type": coll_name,
                "supplier_id": p.get("supplier_id", ""),
                "supplier_name": sup.get("name", ""),
                "quantity": float(p.get("quantity", 0) or 0),
                "rate": float(p.get("purchase_rate", p.get("rate", 0)) or 0),
                "transport": float(p.get("transport", 0) or 0),
                "total": round(total, 2),
            })
    rows.sort(key=lambda r: r["date"], reverse=True)
    total = sum(r["total"] for r in rows)
    return {"rows": rows, "summary": {"count": len(rows), "total": round(total, 2)}}


@api.get("/reports/payments")
async def report_payments(dfrom: Optional[str] = None, dto: Optional[str] = None,
                          customer_id: Optional[str] = None, user=Depends(get_user)):
    q = {"party_type": "customer"}
    if customer_id:
        q = {"$and": [{"$or": [{"customer_id": customer_id}, {"party_id": customer_id}]},
                      {"$or": [{"party_type": "customer"}, {"customer_id": {"$ne": None}}]}]}
    pays = await db.payments.find(q, {"_id": 0}).to_list(20000)
    rows = []
    by_method = {}
    for p in pays:
        if not _within(p, dfrom, dto): continue
        amt = float(p.get("amount", 0) or 0)
        m = (p.get("method") or "cash").lower()
        by_method[m] = round(by_method.get(m, 0) + amt, 2)
        rows.append({
            "id": p.get("id", ""),
            "date": _to_date(p.get("date", "")),
            "customer_id": p.get("customer_id") or p.get("party_id"),
            "customer_name": p.get("party_name", ""),
            "amount": amt,
            "applied_amount": float(p.get("applied_amount", amt) or 0),
            "advance_amount": float(p.get("advance_amount", 0) or 0),
            "method": p.get("method", "cash"),
            "notes": p.get("notes", ""),
            "allocations": p.get("allocations", []) or [],
        })
    rows.sort(key=lambda r: r["date"], reverse=True)
    return {"rows": rows, "summary": {"count": len(rows),
        "total": round(sum(r["amount"] for r in rows), 2),
        "by_method": by_method}}


@api.get("/reports/stock")
async def report_stock(user=Depends(get_user)):
    feed_items = await db.feed_items.find({}, {"_id": 0}).to_list(5000)
    feed_rows = [{
        "id": f["id"], "business_unit": 1, "name": f.get("name", ""),
        "brand": f.get("brand", ""), "unit": f.get("unit", "kg"),
        "current_stock": float(f.get("current_stock", 0) or 0),
        "weighted_avg_cost": float(f.get("weighted_avg_cost", 0) or 0),
        "stock_value": round(float(f.get("current_stock", 0) or 0) * float(f.get("weighted_avg_cost", 0) or 0), 2),
    } for f in feed_items]
    batches = await db.poultry_batches.find({}, {"_id": 0}).to_list(5000)
    batch_rows = [{
        "id": b["id"], "business_unit": 2, "batch_no": b.get("batch_no", ""),
        "hatched_chicks": int(b.get("hatched_chicks", 0) or 0),
        "sold": int(b.get("sold", 0) or 0),
        "transferred": int(b.get("transferred", 0) or 0),
        "available": int(b.get("hatched_chicks", 0) or 0) - int(b.get("sold", 0) or 0) - int(b.get("transferred", 0) or 0),
        "status": b.get("status", ""),
    } for b in batches]
    farm_stock = await db.farm_stock.find({}, {"_id": 0}).to_list(5000)
    farm_rows = [{
        "id": s["id"], "business_unit": 3, "date": s.get("date", ""),
        "current_count": int(s.get("current_count", 0) or 0),
        "source": s.get("source", "transfer"),
    } for s in farm_stock]
    tanks = await db.water_tanks.find({}, {"_id": 0}).to_list(5000)
    water_rows = [{
        "id": t["id"], "business_unit": 4, "name": t.get("name", ""),
        "capacity": float(t.get("capacity", 0) or 0),
        "current_liters": float(t.get("current_liters", 0) or 0),
    } for t in tanks]
    return {
        "feed": feed_rows, "hatchery": batch_rows,
        "farm": farm_rows, "water": water_rows,
        "summary": {
            "feed_value": round(sum(r["stock_value"] for r in feed_rows), 2),
            "available_chicks": sum(r["available"] for r in batch_rows),
            "farm_birds": sum(r["current_count"] for r in farm_rows),
            "water_liters": round(sum(r["current_liters"] for r in water_rows), 2),
        }
    }


@api.get("/reports/bu-summary")
async def report_bu_summary(dfrom: Optional[str] = None, dto: Optional[str] = None,
                            user=Depends(get_user)):
    out = {}
    for coll_name, bu, _ in SALE_COLLECTIONS:
        rows = await db[coll_name].find({}, {"_id": 0}).to_list(20000)
        rev = paid = due = 0.0
        cnt = 0
        for r in rows:
            if not _within(r, dfrom, dto): continue
            rev += float(r.get("total", 0) or 0)
            paid += float(r.get("amount_paid", 0) or 0)
            due += float(r.get("balance_due", 0) or 0)
            cnt += 1
        out[f"bu{bu}"] = {
            "business_unit": bu, "label": BU_LABEL_MAP[bu],
            "sales_count": cnt,
            "revenue": round(rev, 2),
            "collected": round(paid, 2),
            "outstanding": round(due, 2),
        }
    # finance pnl in range
    fin = await db.finance_transactions.find({}, {"_id": 0}).to_list(50000)
    inc = exp = 0.0
    for t in fin:
        if not _within(t, dfrom, dto): continue
        if t.get("type") == "income": inc += float(t.get("amount", 0) or 0)
        elif t.get("type") == "expense": exp += float(t.get("amount", 0) or 0)
    out["combined"] = {"income": round(inc, 2), "expense": round(exp, 2),
                       "profit": round(inc - exp, 2)}
    return out


@api.get("/reports/excel")
async def report_excel(kind: str,
                       dfrom: Optional[str] = None, dto: Optional[str] = None,
                       customer_id: Optional[str] = None,
                       supplier_id: Optional[str] = None,
                       business_unit: Optional[int] = None,
                       user=Depends(get_user)):
    from openpyxl import Workbook
    wb = Workbook()
    ws = wb.active

    def _set_headers(headers):
        ws.append(headers)
        for col in range(1, len(headers) + 1):
            ws.cell(row=1, column=col).font = ws.cell(row=1, column=col).font.copy(bold=True)

    fname = f"{kind}-report.xlsx"
    if kind == "sales":
        ws.title = "Sales"
        data = await report_sales(dfrom, dto, customer_id, business_unit, user=user)
        _set_headers(["Invoice", "Date", "Customer", "BU", "Total", "Paid", "Due", "Status"])
        for r in data["rows"]:
            ws.append([r["invoice_no"], r["date"], r["customer_name"],
                       BU_LABEL_MAP.get(r["business_unit"], r["business_unit"]),
                       r["total"], r["amount_paid"], r["balance_due"], r["payment_status"]])
    elif kind == "purchases":
        ws.title = "Purchases"
        data = await report_purchases(dfrom, dto, supplier_id, business_unit, user=user)
        _set_headers(["Date", "Supplier", "BU", "Qty", "Rate", "Transport", "Total"])
        for r in data["rows"]:
            ws.append([r["date"], r["supplier_name"], BU_LABEL_MAP.get(r["business_unit"], r["business_unit"]),
                       r["quantity"], r["rate"], r["transport"], r["total"]])
    elif kind == "payments":
        ws.title = "Payments"
        data = await report_payments(dfrom, dto, customer_id, user=user)
        _set_headers(["Date", "Customer", "Amount", "Method", "Applied", "Advance", "Notes"])
        for r in data["rows"]:
            ws.append([r["date"], r["customer_name"], r["amount"], r["method"],
                       r["applied_amount"], r["advance_amount"], r["notes"]])
    elif kind == "outstanding":
        ws.title = "Outstanding"
        custs = await db.customers.find({"outstanding": {"$gt": 0}}, {"_id": 0}).to_list(20000)
        _set_headers(["Customer", "Phone", "Credit Limit", "Outstanding"])
        for c in sorted(custs, key=lambda x: x.get("outstanding", 0), reverse=True):
            ws.append([c.get("name", ""), c.get("phone", ""),
                       float(c.get("credit_limit", 0) or 0), float(c.get("outstanding", 0) or 0)])
    elif kind == "stock":
        data = await report_stock(user=user)
        ws.title = "Feed Stock"
        _set_headers(["Name", "Brand", "Unit", "Current", "Avg Cost", "Value"])
        for r in data["feed"]:
            ws.append([r["name"], r["brand"], r["unit"], r["current_stock"],
                       r["weighted_avg_cost"], r["stock_value"]])
        ws2 = wb.create_sheet("Hatchery"); ws2.append(["Batch", "Hatched", "Sold", "Transferred", "Available", "Status"])
        for r in data["hatchery"]:
            ws2.append([r["batch_no"], r["hatched_chicks"], r["sold"], r["transferred"], r["available"], r["status"]])
        ws3 = wb.create_sheet("Farm"); ws3.append(["Date", "Birds", "Source"])
        for r in data["farm"]:
            ws3.append([r["date"], r["current_count"], r["source"]])
        ws4 = wb.create_sheet("Water"); ws4.append(["Tank", "Capacity", "Current"])
        for r in data["water"]:
            ws4.append([r["name"], r["capacity"], r["current_liters"]])
    elif kind == "bu-summary":
        ws.title = "BU Summary"
        data = await report_bu_summary(dfrom, dto, user=user)
        _set_headers(["BU", "Sales Count", "Revenue", "Collected", "Outstanding"])
        for k, v in data.items():
            if k == "combined": continue
            ws.append([v["label"], v["sales_count"], v["revenue"], v["collected"], v["outstanding"]])
    elif kind == "pnl":
        ws.title = "P&L"
        data = await report_bu_summary(dfrom, dto, user=user)
        _set_headers(["Metric", "Value"])
        ws.append(["Income", data["combined"]["income"]])
        ws.append(["Expense", data["combined"]["expense"]])
        ws.append(["Profit", data["combined"]["profit"]])
    else:
        raise HTTPException(400, f"Unknown report kind: {kind}")

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return StreamingResponse(buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{fname}"'})


# ============ Dashboard widgets (top customers, recents) ============
@api.get("/dashboard/top-customers")
async def dashboard_top_customers(by: str = "outstanding", limit: int = 5,
                                  user=Depends(get_user)):
    if by == "revenue":
        # aggregate revenue per customer from sales
        totals = {}
        for coll_name, _bu, _pref in SALE_COLLECTIONS:
            rows = await db[coll_name].find({}, {"customer_id": 1, "customer_name": 1, "total": 1, "_id": 0}).to_list(50000)
            for r in rows:
                cid = r.get("customer_id")
                if not cid: continue
                d = totals.setdefault(cid, {"id": cid, "name": r.get("customer_name", ""), "revenue": 0.0})
                d["revenue"] += float(r.get("total", 0) or 0)
        ranked = sorted(totals.values(), key=lambda x: x["revenue"], reverse=True)[:limit]
        for r in ranked: r["revenue"] = round(r["revenue"], 2)
        return ranked
    # default by outstanding
    custs = await db.customers.find({"outstanding": {"$gt": 0}}, {"_id": 0}).sort("outstanding", -1).limit(limit).to_list(limit)
    return [{"id": c["id"], "name": c.get("name", ""),
             "phone": c.get("phone", ""),
             "outstanding": float(c.get("outstanding", 0) or 0)} for c in custs]


@api.get("/dashboard/recent-payments")
async def dashboard_recent_payments(limit: int = 10, user=Depends(get_user)):
    rows = await db.payments.find({}, {"_id": 0}).sort("created_at", -1).limit(limit).to_list(limit)
    return rows


@api.get("/dashboard/recent-sales")
async def dashboard_recent_sales(limit: int = 10, user=Depends(get_user)):
    combined = []
    for coll_name, bu, _ in SALE_COLLECTIONS:
        rows = await db[coll_name].find({}, {"_id": 0}).sort("created_at", -1).limit(limit).to_list(limit)
        for r in rows:
            combined.append({
                "id": r["id"], "invoice_no": r.get("invoice_no", ""),
                "date": _to_date(r.get("date", "")),
                "created_at": r.get("created_at", ""),
                "customer_name": r.get("customer_name", ""),
                "business_unit": bu, "sale_type": coll_name,
                "total": float(r.get("total", 0) or 0),
                "payment_status": r.get("payment_status", "pending"),
            })
    combined.sort(key=lambda r: r.get("created_at", ""), reverse=True)
    return combined[:limit]


# ============ Executive Dashboard (Reports) ============
@api.get("/reports/exec-dashboard")
async def reports_exec_dashboard(dfrom: Optional[str] = None, dto: Optional[str] = None,
                                 user=Depends(get_user)):
    # Per-BU sales (revenue, outstanding, count)
    per_bu = {1: {"revenue": 0.0, "outstanding": 0.0, "sales_count": 0},
              2: {"revenue": 0.0, "outstanding": 0.0, "sales_count": 0},
              3: {"revenue": 0.0, "outstanding": 0.0, "sales_count": 0},
              4: {"revenue": 0.0, "outstanding": 0.0, "sales_count": 0}}
    for coll_name, bu, _ in SALE_COLLECTIONS:
        rows = await db[coll_name].find({}, {"_id": 0}).to_list(50000)
        for r in rows:
            if not _within(r, dfrom, dto): continue
            per_bu[bu]["revenue"] += float(r.get("total", 0) or 0)
            per_bu[bu]["outstanding"] += float(r.get("balance_due", 0) or 0)
            per_bu[bu]["sales_count"] += 1

    # Finance expenses & income per BU (within range)
    fin_per_bu = {1: {"income": 0.0, "expense": 0.0},
                  2: {"income": 0.0, "expense": 0.0},
                  3: {"income": 0.0, "expense": 0.0},
                  4: {"income": 0.0, "expense": 0.0}}
    expense_by_category = {}
    fin = await db.finance_transactions.find({}, {"_id": 0}).to_list(100000)
    for t in fin:
        if not _within(t, dfrom, dto): continue
        bu = int(t.get("business_unit", 0) or 0)
        amt = float(t.get("amount", 0) or 0)
        typ = t.get("type", "")
        if bu in fin_per_bu:
            if typ == "income": fin_per_bu[bu]["income"] += amt
            elif typ == "expense": fin_per_bu[bu]["expense"] += amt
        if typ == "expense":
            cat = t.get("category", "Other") or "Other"
            expense_by_category[cat] = expense_by_category.get(cat, 0.0) + amt

    # Stock value per BU
    stock_value_bu = {1: 0.0, 2: 0.0, 3: 0.0, 4: 0.0}
    stock_units_bu = {1: 0.0, 2: 0.0, 3: 0.0, 4: 0.0}
    feed_items = await db.feed_items.find({}, {"_id": 0}).to_list(5000)
    for f in feed_items:
        qty = float(f.get("current_stock", 0) or 0)
        cost = float(f.get("weighted_avg_cost", 0) or 0)
        stock_value_bu[1] += qty * cost
        stock_units_bu[1] += qty
    batches = await db.poultry_batches.find({}, {"_id": 0}).to_list(5000)
    for b in batches:
        avail = int(b.get("hatched_chicks", 0) or 0) - int(b.get("sold", 0) or 0) - int(b.get("transferred", 0) or 0)
        if avail < 0: avail = 0
        stock_units_bu[2] += avail
    farm_stock = await db.farm_stock.find({}, {"_id": 0}).to_list(5000)
    for s in farm_stock:
        stock_units_bu[3] += int(s.get("current_count", 0) or 0)
    tanks = await db.water_tanks.find({}, {"_id": 0}).to_list(5000)
    for t in tanks:
        stock_units_bu[4] += float(t.get("current_liters", 0) or 0)

    # Compose per-BU result
    BU_UNITS = {1: "kg", 2: "chicks", 3: "birds", 4: "L"}
    bu_result = {}
    for bu in (1, 2, 3, 4):
        rev = round(per_bu[bu]["revenue"], 2)
        exp = round(fin_per_bu[bu]["expense"], 2)
        bu_result[f"bu{bu}"] = {
            "business_unit": bu,
            "label": BU_LABEL_MAP[bu],
            "revenue": rev,
            "expenses": exp,
            "profit": round(rev - exp, 2),
            "outstanding": round(per_bu[bu]["outstanding"], 2),
            "stock_value": round(stock_value_bu[bu], 2),
            "stock_units": round(stock_units_bu[bu], 2),
            "stock_unit_label": BU_UNITS[bu],
            "sales_count": per_bu[bu]["sales_count"],
        }

    total_rev = round(sum(v["revenue"] for v in bu_result.values()), 2)
    total_exp = round(sum(v["expenses"] for v in bu_result.values()), 2)
    total_out = round(sum(v["outstanding"] for v in bu_result.values()), 2)
    total_stk = round(sum(v["stock_value"] for v in bu_result.values()), 2)

    # Monthly trend (last 6 months ending today or dto)
    from datetime import datetime as _dt
    end_ref = (dto or _dt.utcnow().strftime("%Y-%m-%d"))[:7]  # YYYY-MM
    try:
        year, month = int(end_ref[:4]), int(end_ref[5:7])
    except Exception:
        now = _dt.utcnow(); year, month = now.year, now.month
    months = []
    y, m = year, month
    for _ in range(6):
        months.append(f"{y:04d}-{m:02d}")
        m -= 1
        if m == 0:
            m = 12; y -= 1
    months.reverse()
    trend_map = {mo: {"month": mo, "revenue": 0.0, "expenses": 0.0, "profit": 0.0} for mo in months}
    # Revenue from sales (consistent with totals.revenue)
    for coll_name, _bu, _ in SALE_COLLECTIONS:
        rows = await db[coll_name].find({}, {"_id": 0}).to_list(50000)
        for r in rows:
            d = (r.get("date") or "")[:7]
            if d in trend_map:
                trend_map[d]["revenue"] += float(r.get("total", 0) or 0)
    # Expenses from finance (only true outflows)
    for t in fin:
        if t.get("type") != "expense": continue
        d = (t.get("date") or "")[:7]
        if d in trend_map:
            trend_map[d]["expenses"] += float(t.get("amount", 0) or 0)
    monthly_trend = []
    for mo in months:
        r = trend_map[mo]
        r["revenue"] = round(r["revenue"], 2)
        r["expenses"] = round(r["expenses"], 2)
        r["profit"] = round(r["revenue"] - r["expenses"], 2)
        monthly_trend.append(r)

    expense_breakdown = [{"category": k, "amount": round(v, 2)}
                         for k, v in sorted(expense_by_category.items(), key=lambda x: -x[1])]

    return {
        "from": dfrom, "to": dto,
        "totals": {
            "revenue": total_rev, "expenses": total_exp,
            "profit": round(total_rev - total_exp, 2),
            "outstanding": total_out, "stock_value": total_stk,
        },
        "per_bu": bu_result,
        "monthly_trend": monthly_trend,
        "expense_breakdown": expense_breakdown,
    }


# ============ Startup ============
app.include_router(api)
app.add_middleware(CORSMiddleware, allow_credentials=True,
                   allow_origins=[o.strip() for o in os.environ.get("CORS_ORIGINS","").split(",") if o.strip()],
                   allow_methods=["*"], allow_headers=["*"])

logging.basicConfig(level=logging.INFO)
logger = logging.getLogger(__name__)

@app.on_event("startup")
async def startup():
    for coll in ["users","customers","suppliers","feed_items","feed_purchases","feed_sales",
                 "egg_purchases","poultry_batches","batch_expenses","chick_sales",
                 "farm_stock","farm_sales","farm_expenses","water_tanks","water_tank_additions",
                 "water_sales","water_expenses","internal_transfers","payments","finance_transactions"]:
        await db[coll].create_index("id", unique=True, sparse=True)
    await db.users.create_index("email", unique=True)
    await db.finance_transactions.create_index([("date", 1), ("business_unit", 1)])

    admin_email = os.environ.get("ADMIN_EMAIL", "admin@agribiz.com")
    admin_password = os.environ.get("ADMIN_PASSWORD", "admin123")
    existing = await db.users.find_one({"email": admin_email})
    if not existing:
        await db.users.insert_one({"id": new_id(), "email": admin_email,
            "password_hash": hp(admin_password), "name": "Admin", "role": "admin",
            "created_at": now_iso()})
        logger.info(f"Seeded admin: {admin_email}")
    elif not vp(admin_password, existing["password_hash"]):
        await db.users.update_one({"email": admin_email}, {"$set": {"password_hash": hp(admin_password)}})

    # Migrate legacy sales: ensure amount_paid / balance_due / business_unit / payment_status exist
    try:
        for coll_name, bu, _pref in SALE_COLLECTIONS:
            await _ensure_sale_payment_fields(coll_name, bu)
        # Recompute outstanding for every customer from balance_due totals
        customers_list = await db.customers.find({}, {"id": 1, "_id": 0}).to_list(20000)
        for c in customers_list:
            await _recompute_customer_outstanding(c["id"])
        logger.info("Payment-fields migration complete")
    except Exception as e:
        logger.warning(f"Migration warning: {e}")
